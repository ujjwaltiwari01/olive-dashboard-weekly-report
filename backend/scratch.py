import openpyxl

from excel_parser import EXCEL_PATH

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for c_idx, cell in enumerate(row, 1):
            if isinstance(cell, str) and 'Target' in cell:
                print(f'Found "{cell}" in {sheet_name} at Row {r_idx}, Col {c_idx}')
                
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for c_idx, cell in enumerate(row, 1):
            if isinstance(cell, str) and 'Actual' in cell:
                print(f'Found "{cell}" in {sheet_name} at Row {r_idx}, Col {c_idx}')
