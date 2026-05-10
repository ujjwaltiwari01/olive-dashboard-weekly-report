"""
Excel Parser — reads the CFO weekly workbook
`Weekly update - 08.05.2026 v1.xlsx` (multi-sheet).

Production (Railway, Docker): set OLIVE_WEEKLY_EXCEL_PATH or EXCEL_PATH to an absolute path
for this exact workbook name. If unset, the file next to the repo root is used:
Weekly update - 08.05.2026 v1.xlsx
"""
import os
import threading
import openpyxl
import pandas as pd

# Canonical weekly workbook for this dashboard (repo root). All KPI readers use EXCEL_PATH.
# Override only via OLIVE_WEEKLY_EXCEL_PATH / EXCEL_PATH — do not hardcode alternate .xlsx names in KPI modules.
WEEKLY_WORKBOOK_FILENAME = "Weekly update - 08.05.2026 v1.xlsx"
_DEFAULT_WORKBOOK = WEEKLY_WORKBOOK_FILENAME


def resolve_excel_path() -> str:
    """Resolve the canonical workbook path.

    Default: `<repo-root>/Weekly update - 08.05.2026 v1.xlsx`.
    Env overrides (`OLIVE_WEEKLY_EXCEL_PATH`, `EXCEL_PATH`) are honoured **only** if they
    point at this exact workbook name. Any stale path to an older weekly workbook is
    ignored, so the dashboard cannot silently load the wrong data.
    """
    here = os.path.dirname(os.path.abspath(__file__))
    default_path = os.path.abspath(os.path.join(here, "..", _DEFAULT_WORKBOOK))

    for key in ("OLIVE_WEEKLY_EXCEL_PATH", "EXCEL_PATH"):
        raw = os.environ.get(key, "").strip().strip('"').strip("'")
        if not raw:
            continue
        candidate = os.path.abspath(os.path.expanduser(raw))
        if os.path.isfile(candidate) and os.path.basename(candidate) == WEEKLY_WORKBOOK_FILENAME:
            return candidate
    return default_path


EXCEL_PATH = resolve_excel_path()


def excel_source_path() -> str:
    return EXCEL_PATH


def excel_file_available() -> bool:
    return os.path.isfile(EXCEL_PATH)


def excel_workbook_missing_message() -> str:
    """Single user-facing hint when the weekly xlsx is absent (API / exceptions)."""
    return (
        "Weekly Excel workbook is missing. "
        f"Add `{WEEKLY_WORKBOOK_FILENAME}` at the repository root (beside `backend/` and `frontend/`), "
        "or set OLIVE_WEEKLY_EXCEL_PATH / EXCEL_PATH to its absolute path (e.g. Railway Variables). "
        f"Resolved path: {excel_source_path()}"
    )


# One parsed workbook in memory per file mtime — avoids reopening the .xlsx on every KPI call
# (previously each `get_sheet_values` did a full `load_workbook`, which is very slow).
_wb_lock = threading.RLock()
_workbook_mtime: float | None = None
_workbook: openpyxl.Workbook | None = None
_sheet_rows_cache: dict[str, list[list]] = {}


def shared_workbook() -> openpyxl.Workbook:
    """Return the shared workbook (data_only). Reloads only when `EXCEL_PATH` file mtime changes.

    Callers must not call `.close()` on the returned object; the cache owns the lifecycle.
    """
    if not excel_file_available():
        raise FileNotFoundError(excel_workbook_missing_message())
    with _wb_lock:
        _ensure_workbook_loaded()
        assert _workbook is not None
        return _workbook


def _ensure_workbook_loaded() -> None:
    global _workbook_mtime, _workbook, _sheet_rows_cache
    with _wb_lock:
        try:
            mtime = os.path.getmtime(EXCEL_PATH)
        except OSError:
            mtime = 0.0
        if _workbook is not None and _workbook_mtime == mtime:
            return
        if _workbook is not None:
            try:
                _workbook.close()
            except Exception:
                pass
            _workbook = None
        _sheet_rows_cache.clear()
        _workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        _workbook_mtime = mtime


def get_sheet_values(sheet_name: str) -> list[list]:
    """Return all rows (as list of values) from a given sheet (cached per sheet until file changes)."""
    if not excel_file_available():
        return []
    _ensure_workbook_loaded()
    assert _workbook is not None
    with _wb_lock:
        if sheet_name in _sheet_rows_cache:
            return _sheet_rows_cache[sheet_name]
        if sheet_name not in _workbook.sheetnames:
            _sheet_rows_cache[sheet_name] = []
            return []
        ws = _workbook[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        _sheet_rows_cache[sheet_name] = rows
        return rows


def safe_float(val) -> float | None:
    """Parse a cell value as float. Handles Indian-style commas (e.g. 1,79,56,000.00), ₹, %, and accounting negatives."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        try:
            f = float(val)
        except (TypeError, ValueError):
            return None
        if f != f:  # NaN
            return None
        return f
    if isinstance(val, str):
        s = val.strip()
        if not s or s.upper() in ("#N/A", "N/A", "-", "–", "—"):
            return None
        if s.startswith("#") and len(s) > 1:
            return None
        pct_suffix = s.endswith("%")
        if pct_suffix:
            s = s[:-1].strip()
        s = s.replace("₹", "").replace(",", "").replace(" ", "")
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()
        if not s:
            return None
        try:
            out = float(s)
        except ValueError:
            return None
        if neg:
            out = -out
        return out
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def safe_int(val) -> int | None:
    f = safe_float(val)
    return int(f) if f is not None else None
