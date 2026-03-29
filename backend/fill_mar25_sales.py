import openpyxl
import os
import random

FILE_PATH = os.path.join(os.path.dirname(__file__), "..", "19.Olive-Weekly Status update_23-Mar-2026.xlsx")

def fill_mar25_data():
    if not os.path.exists(FILE_PATH):
        print(f"Error: {FILE_PATH} not found.")
        return

    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    
    if "Sales" not in wb.sheetnames:
        print("Error: Sales sheet not found.")
        return
        
    ws = wb["Sales"]
    
    # We want to use columns 61, 62, 63, 64 (1-indexed) in Excel.
    # 61: Corporate, 62: Walkins, 63: Online, 64: TOTAL.
    col_corp = 61
    col_walk = 62
    col_onli = 63
    col_tot  = 64
    
    # Write Headers
    ws.cell(row=1, column=col_corp, value="Mar'25")
    ws.cell(row=2, column=col_corp, value="Corporate")
    ws.cell(row=2, column=col_walk, value="Walkins")
    ws.cell(row=2, column=col_onli, value="Online")
    ws.cell(row=2, column=col_tot, value="TOTAL")
    
    print("Writing dummy entries for March 2025 (YoY Comparison)...")
    
    # Target totals for Mar'25 (per spec): Walk-in: ~10%, Online: ~10%, Corporate: ~80%
    # Overall we need around ~15% less than March 2026. 
    # M26 Total was ~781 Lakhs. M25 Should be around 680 Lakhs. 
    # Walkin: 68L (6.8m), Online: 68L (6.8m), Corporate: 544L (54.4m)
    
    # We will distribute these amounts randomly across all properties (rows 3 to ~60).
    
    corp_sum = 54400000
    walk_sum = 6800000
    onli_sum = 6800000
    
    properties = []
    # Identify property rows
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        # Check if cell has text and doesn't start with TOTAL
        if val and isinstance(val, str) and not val.strip().upper().startswith("TOTAL"):
            properties.append(r)
            
    num_props = len(properties)
    
    if num_props > 0:
        # Give roughly equal amount plus some randomness
        for i, row_idx in enumerate(properties):
            c = corp_sum // num_props + random.randint(-50000, 50000)
            w = walk_sum // num_props + random.randint(-10000, 10000)
            o = onli_sum // num_props + random.randint(-10000, 10000)
            
            c = max(0, c)
            w = max(0, w)
            o = max(0, o)
            
            ws.cell(row=row_idx, column=col_corp, value=c)
            ws.cell(row=row_idx, column=col_walk, value=w)
            ws.cell(row=row_idx, column=col_onli, value=o)
            ws.cell(row=row_idx, column=col_tot, value=c+w+o)
            
    print("Saving workbook... this may take a few seconds.")
    wb.save(FILE_PATH)
    print("Done! Dummy data for March '25 appended successfully.")

if __name__ == "__main__":
    fill_mar25_data()
