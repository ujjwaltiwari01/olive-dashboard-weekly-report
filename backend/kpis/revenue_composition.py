from __future__ import annotations

"""
KPI: Revenue Composition — reads the Revenue sheet from the workbook at
`excel_parser.EXCEL_PATH`. By default that file is **`Weekly update - 30.04.2024 v3.xlsx`**
at the repository root (one level above `backend/`). Override only with
`OLIVE_WEEKLY_EXCEL_PATH` or `EXCEL_PATH` if the workbook lives elsewhere.

SECTION 1: Target vs Actual (managed)     → rows 5,6,8  | achievement col J (10)
SECTION 2: YoY managed (Short / Long stay) → rows 14,15,17 | growth col J (10)
SECTION 3: MoM franchised — anchored on title "MoM Growth - Franchised…" then
            month labels row +2, values Open/Others/Total at +4,+5,+7 (same columns as labels).
"""
import os
import openpyxl

from excel_parser import EXCEL_PATH, excel_workbook_missing_message, safe_float


def _read_revenue_sheet():
    if not os.path.isfile(EXCEL_PATH):
        raise FileNotFoundError(excel_workbook_missing_message())
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Revenue"]

    def val(row, col):
        v = ws.cell(row, col).value
        f = safe_float(v)
        return f if f is not None else 0.0

    def pct(row, col):
        """Read a % cell (stored as decimal in Excel, or as '45%' / text) and return as integer %."""
        v = ws.cell(row, col).value
        if v is None:
            return 0
        f = safe_float(v)
        if f is None:
            return 0
        if isinstance(v, str) and "%" in v:
            return round(f)
        return round(f * 100)

    # ── SECTION 1: Target vs Actual (managed — Online / Offline) ─────────────
    target_online = val(5, 3)
    target_offline = val(6, 3)
    target_total = val(8, 3)

    actual_online = val(5, 6)
    actual_offline = val(6, 6)
    actual_total = val(8, 6)

    achievement_pct = round(actual_total / target_total * 100) if target_total else 0
    achievement_online_pct = pct(5, 10)
    achievement_offline_pct = pct(6, 10)

    section1 = {
        "label": "Target vs Actual (April'26)",
        "achievement_pct": achievement_pct,
        "achievement_online_pct": achievement_online_pct,
        "achievement_offline_pct": achievement_offline_pct,
        "bars": [
            {
                "name": "Target",
                "online": round(target_online),
                "offline": round(target_offline),
                "total": round(target_total),
                "online_pct": round(target_online / target_total * 100) if target_total else 0,
                "offline_pct": round(target_offline / target_total * 100) if target_total else 0,
            },
            {
                "name": "Actual",
                "online": round(actual_online),
                "offline": round(actual_offline),
                "total": round(actual_total),
                "online_pct": round(actual_online / actual_total * 100) if actual_total else 0,
                "offline_pct": round(actual_offline / actual_total * 100) if actual_total else 0,
            },
        ],
    }

    # ── SECTION 2: YoY — Short stay / Long stay (cols C & F, totals row 17) ─
    s25 = val(14, 3)
    l25 = val(15, 3)
    t25 = val(17, 3)

    s26 = val(14, 6)
    l26 = val(15, 6)
    t26 = val(17, 6)

    yoy_pct = round((t26 - t25) / t25 * 100) if t25 else 0
    yoy_short_stay_pct = pct(14, 10)
    yoy_long_stay_pct = pct(15, 10)

    section2 = {
        "label": "YoY Growth (April 2025 vs April 2026)",
        "yoy_pct": yoy_pct,
        "yoy_short_stay_pct": yoy_short_stay_pct,
        "yoy_long_stay_pct": yoy_long_stay_pct,
        "bars": [
            {
                "name": "April 2025",
                "short_stay": round(s25),
                "long_stay": round(l25),
                "total": round(t25),
                "short_stay_pct": round(s25 / t25 * 100) if t25 else 0,
                "long_stay_pct": round(l25 / t25 * 100) if t25 else 0,
            },
            {
                "name": "April 2026",
                "short_stay": round(s26),
                "long_stay": round(l26),
                "total": round(t26),
                "short_stay_pct": round(s26 / t26 * 100) if t26 else 0,
                "long_stay_pct": round(l26 / t26 * 100) if t26 else 0,
            },
        ],
    }

    # ── SECTION 3: MoM franchised (Open vs others), month columns under sheet header ─
    # v2 layout: title ~R19 "MoM Growth - Franchised properties", month labels R21 (C,E,G),
    # Open R23, others R24, Total R26 — same column as each month label.
    # Do not rely on fixed rows alone: if EXCEL_PATH points at another workbook revision,
    # blind (23,3)/(23,5)/(23,7) reads pull unrelated cells. Anchor on the title row instead.

    def franchised_bar(
        name: str, o: float, ot: float, tot: float, contrib: float | None
    ) -> dict:
        t = float(tot) if tot and tot > 0 else float(o) + float(ot)
        if t <= 0:
            t = max(float(o) + float(ot), 1e-9)
        contrib_pct = None
        if contrib is not None and contrib >= 0:
            if contrib <= 1.0:
                contrib_pct = round(contrib * 100, 2)
            else:
                contrib_pct = round(contrib, 2)
        out = {
            "name": name,
            "fr_open": round(o, 2),
            "fr_others": round(ot, 2),
            "total": round(t, 2),
            "fr_open_pct": round(o / t * 100) if t else 0,
            "fr_others_pct": round(ot / t * 100) if t else 0,
        }
        if contrib_pct is not None:
            out["contribution_pct"] = contrib_pct
        return out

    def _find_franchised_title_row(ws) -> int | None:
        for r in range(1, 85):
            for c in range(1, 6):
                t = str(ws.cell(r, c).value or "").strip().lower()
                if "franchis" in t and "mom" in t:
                    return r
        return None

    def _scan_franchised_value_rows(ws, title_row: int) -> tuple[int, int | None, int | None, int | None]:
        """Locate Open / Others / Total / % of contribution rows by col-B label (v3 has no Others/Total; row 26 is contribution only)."""
        open_r: int | None = None
        others_r: int | None = None
        total_r: int | None = None
        contrib_r: int | None = None
        for r in range(title_row + 2, min(title_row + 16, 90)):
            lab = str(ws.cell(r, 2).value or "").strip().lower()
            if not lab:
                continue
            if "contribution" in lab:
                contrib_r = r
                continue
            if open_r is None and "open" in lab and "other" not in lab:
                open_r = r
                continue
            if others_r is None and "other" in lab:
                others_r = r
                continue
            if total_r is None and "total" in lab and "contribution" not in lab:
                total_r = r
                continue
        if open_r is None:
            return None, None, None, None
        return open_r, others_r, total_r, contrib_r

    def _read_franchised_bars(ws) -> list[dict]:
        """Read Feb / March / April from columns C,E,G (3,5,7). Totals follow a dedicated Total row when present; else Open+Others (v3 workbook has only Open + % of contribution)."""
        title_row = _find_franchised_title_row(ws)
        if title_row is not None:
            open_r, others_r, total_r, contrib_r = _scan_franchised_value_rows(ws, title_row)
            if open_r is None:
                return []
            open_label = str(ws.cell(open_r, 2).value or "").strip().lower()
            if "open" not in open_label:
                return []
        else:
            c21 = str(ws.cell(21, 3).value or "").strip().lower()
            if not c21.startswith("feb"):
                return []
            open_r, others_r, total_r, contrib_r = _scan_franchised_value_rows(ws, 19)
            if open_r is None:
                return []
            open_label = str(ws.cell(open_r, 2).value or "").strip().lower()
            if "open" not in open_label:
                return []

        # Canonical month columns (1-based): C, E, G — display names fixed for the chart.
        month_triple: list[tuple[int, str]] = [
            (3, "Feb"),
            (5, "March"),
            (7, "April"),
        ]

        bars: list[dict] = []
        for col, name in month_triple:
            o = val(open_r, col)
            ot = val(others_r, col) if others_r else 0.0
            sheet_tot = val(total_r, col) if total_r else 0.0
            tot = sheet_tot if sheet_tot > 0 else o + ot
            # Guard: never treat % of contribution decimals (e.g. 0.028) as rupee totals.
            if tot < 200 and max(o, ot) > 2000:
                tot = o + ot
            contrib = float(val(contrib_r, col)) if contrib_r else None
            if tot <= 0 and o <= 0 and ot <= 0:
                continue
            bars.append(franchised_bar(name, o, ot, tot, contrib))

        return bars

    bars_fr = _read_franchised_bars(ws)
    if len(bars_fr) >= 2:
        prev = bars_fr[-2]
        last = bars_fr[-1]
        po, lo = float(prev["fr_open"]), float(last["fr_open"])
        pto, lto = float(prev["fr_others"]), float(last["fr_others"])
        pt, lt = float(prev["total"]), float(last["total"])
        mom_open = round((lo - po) / po * 100, 1) if po else 0.0
        mom_others = round((lto - pto) / pto * 100, 1) if pto else 0.0
        mom_total = round((lt - pt) / pt * 100, 1) if pt else 0.0
    else:
        mom_open = mom_others = mom_total = 0.0

    section3 = {
        "label": "MoM Growth - Franchised properties",
        "mom_open_pct": mom_open,
        "mom_others_pct": mom_others,
        "mom_total_pct": mom_total,
        "bars": bars_fr,
    }

    wb.close()
    return section1, section2, section3


def get_revenue_composition() -> dict:
    try:
        s1, s2, s3 = _read_revenue_sheet()
    except Exception as e:
        return {"error": str(e)}

    return {
        "section1": s1,
        "section2": s2,
        "section3": s3,
    }
