"""
KPI 4: Cashflow — March
Reads live from Excel (cashflow sheet), Weekly update - 27.04.2024 v2.xlsx

cashflow sheet layout (1-indexed cols):
  col2 = label | col3 = Target | col4..7 = W1..W4 | col8 = Total | col9 = Expected
  R8  : TA Fees
  R9  : Management Fees
  R10 : Profit Incentive
  R12 : Outflow total (col9 fallback col8)
  R15 : Current Bank Balance (col9 fallback col8)
  R16 : Coliving Operations (col3)
  R17 : Coliving (col3)
  R18 : VARS (col3)
  R20 : Closing Cash (col9 fallback col8)
"""
import os
import openpyxl

from excel_parser import EXCEL_PATH, excel_workbook_missing_message


def _read():
    if not os.path.isfile(EXCEL_PATH):
        raise FileNotFoundError(excel_workbook_missing_message())
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["cashflow"]

    def v(row, col):
        val = ws.cell(row, col).value
        return float(val) if val is not None else 0.0

    def v_first(row, cols: list[int]) -> float:
        for col in cols:
            val = ws.cell(row, col).value
            if val is not None:
                try:
                    return float(val)
                except (TypeError, ValueError):
                    continue
        return 0.0

    def inflow_row(row: int, label: str) -> dict:
        w1, w2, w3, w4 = v(row, 4), v(row, 5), v(row, 6), v(row, 7)
        total_received = v_first(row, [8]) or (w1 + w2 + w3 + w4)
        expected = v_first(row, [9, 8])
        return {
            "type": label,
            "target": v(row, 3),
            "w1": w1,
            "w2": w2,
            "w3": w3,
            "w4": w4,
            "total_received": total_received,
            "expected": expected,
        }

    inflows = [
        inflow_row(8, "TA Fees"),
        inflow_row(9, "Management Fees"),
        inflow_row(10, "Profit Incentive"),
    ]

    for inf in inflows:
        inf["received"] = inf["total_received"]

    total_outflow    = v_first(12, [9, 8])   # Outflow row
    current_balance  = v_first(15, [9, 8])   # Current bank balance
    closing_balance  = v_first(20, [9, 8])   # Closing Cash

    account_balance = [
        {"account": "Coliving Operations", "amount": v(16, 3)},
        {"account": "Coliving",            "amount": v(17, 3)},
        {"account": "VARS",                "amount": v(18, 3)},
    ]

    wb.close()
    return inflows, total_outflow, current_balance, closing_balance, account_balance


def get_cashflow() -> dict:
    try:
        inflows, total_outflow, current_balance, closing_balance, account_balance = _read()
    except Exception as e:
        return {"error": str(e)}

    return {
        "inflows":           inflows,
        "account_balance":   account_balance,
        "inflow_totals": {
            "target":   sum(i['target']   for i in inflows),
            "received": sum(i['received'] for i in inflows),
            "w1":       sum(i['w1']       for i in inflows),
            "w2":       sum(i['w2']       for i in inflows),
            "w3":       sum(i['w3']       for i in inflows),
            "w4":       sum(i['w4']       for i in inflows),
            "total_received": sum(i['total_received'] for i in inflows),
            "expected": sum(i['expected'] for i in inflows),
        },
        "summary": {
            "immediate_payments": total_outflow,
            "current_balance":    current_balance,
            "total_inflow":       sum(i['received'] for i in inflows),
            "total_outflow":      total_outflow,
            "closing_balance":    closing_balance,
        },
    }
