import openpyxl
import os
import datetime

FILE_PATH = os.path.join(os.path.dirname(__file__), "..", "19.Olive-Weekly Status update_23-Mar-2026.xlsx")

def fill_mar25_revenue():
    if not os.path.exists(FILE_PATH):
        print(f"Error: {FILE_PATH} not found.")
        return

    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    
    if "Summary" not in wb.sheetnames:
        print("Error: Summary sheet not found.")
        return
        
    ws = wb["Summary"]
    
    # Target columns for Mar'25 (Col 40, 41, 42)
    # 40: Target, 41: Achieved, 42: Achieved %
    
    col_tgt = 40
    col_ach = 41
    col_pct = 42
    
    # Write Row 3 (Headers - Dates)
    ws.cell(row=3, column=col_tgt, value=datetime.datetime(2025, 3, 1))
    
    # Write Row 4 (Sub-Headers)
    ws.cell(row=4, column=col_tgt, value="Target")
    ws.cell(row=4, column=col_ach, value="Achieved")
    ws.cell(row=4, column=col_pct, value="Achieved %")
    
    print("Writing dummy entries for March 2025 Revenue Streams (YoY Comparison)...")
    
    # Rows for data (1-indexed based on exact native indices):
    # Row 5 (TA Spark)
    # Row 6 (TA Olive)
    # Row 9 (Mgt Olive)
    # Row 12 (Profit Vendors)
    
    # Achieved mock values (Col 41)
    ws.cell(row=6, column=col_ach, value=4000000)   # Spark TA (~40L)
    ws.cell(row=7, column=col_ach, value=2000000)   # Olive TA (~20L)
    ws.cell(row=10, column=col_ach, value=10000000) # Mgt Fee  (~100L)
    ws.cell(row=13, column=col_ach, value=300000)   # Profit Inc(~3L)

    print("Saving workbook... this may take a few seconds.")
    wb.save(FILE_PATH)
    print("Done! Dummy data for March '25 appended successfully.")

if __name__ == "__main__":
    fill_mar25_revenue()
