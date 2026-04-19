import openpyxl

from excel_parser import EXCEL_PATH

wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb['Signings']

# Find the next empty row, typically around 200
start_row = 200

headers = ['Brand', 'Month', 'Target', 'W1', 'W2', 'W3', 'W4']
for col, header in enumerate(headers, start=1):
    sheet.cell(row=start_row, column=col, value=header)

# Data for KPI 1 Part B (March)
# As per kpi1.md
dummy_data = [
    # Oct-25
    ['Olive', 'Oct-25', 5, 1, 1, 1, 1],
    ['Spark', 'Oct-25', 2, 0, 1, 0, 0],
    ['Open', 'Oct-25', 3, 1, 0, 1, 0],
    # Nov-25
    ['Olive', 'Nov-25', 6, 1, 1, 2, 1],
    ['Spark', 'Nov-25', 2, 0, 1, 0, 0],
    ['Open', 'Nov-25', 4, 1, 1, 1, 0],
    # Dec-25
    ['Olive', 'Dec-25', 5, 1, 1, 1, 1],
    ['Spark', 'Dec-25', 3, 1, 1, 0, 0],
    ['Open', 'Dec-25', 5, 1, 1, 1, 1],
    # Jan-26
    ['Olive', 'Jan-26', 6, 2, 1, 1, 1],
    ['Spark', 'Jan-26', 2, 1, 0, 0, 0],
    ['Open', 'Jan-26', 4, 1, 1, 1, 0],
    # Feb-26
    ['Olive', 'Feb-26', 7, 2, 1, 2, 1],
    ['Spark', 'Feb-26', 1, 0, 0, 0, 0],
    ['Open', 'Feb-26', 6, 2, 1, 1, 1],
    # Mar-26 (Mock data from user)
    ['Olive', 'Mar-26', 7, 2, 1, 1, 2],
    ['Spark', 'Mar-26', 0, 0, 0, 0, 0],
    ['Open', 'Mar-26', 9, 1, 4, 1, 1],
]

for row_idx, row_data in enumerate(dummy_data, start=start_row + 1):
    for col_idx, val in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=val)

wb.save(EXCEL_PATH)
print(f"Dummy data inserted from row {start_row} to {start_row + len(dummy_data)}")
