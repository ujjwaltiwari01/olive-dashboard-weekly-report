import openpyxl

from excel_parser import EXCEL_PATH

wb = openpyxl.load_workbook(EXCEL_PATH)

# 1. TA FEE - Inject W1, W2, W3, W4, Target, Expected, Collected Outstanding
ta_sheet = wb["TA Fee"]
start_col = 41

headers = ["Collected Outstanding", "Target", "W1", "W2", "W3", "W4", "Expected"]
for i, h in enumerate(headers):
    ta_sheet.cell(row=1, column=start_col + i).value = "KPI7"
    ta_sheet.cell(row=2, column=start_col + i).value = h

# Give data to a few rows so it renders well
mock_ta = [
    {"row": 3, "out": 150000, "tgt": 200000, "w1": 10000, "w2": 20000, "w3": 15000, "w4": 5000, "exp": 100000},
    {"row": 4, "out": 400000, "tgt": 500000, "w1": 0, "w2": 0, "w3": 50000, "w4": 50000, "exp": 300000},
    {"row": 5, "out": 50000, "tgt": 80000, "w1": 10000, "w2": 10000, "w3": 10000, "w4": 0, "exp": 40000},
]
# Let's write an OLIVE header further down just to ensure property grouping
ta_sheet.cell(row=20, column=2).value = "OLIVE"
mock_ta.extend([
    {"row": 21, "out": 200000, "tgt": 300000, "w1": 50000, "w2": 25000, "w3": 25000, "w4": 0, "exp": 150000},
    {"row": 22, "out": 800000, "tgt": 1000000, "w1": 100000, "w2": 0, "w3": 0, "w4": 100000, "exp": 500000},
])

for m in mock_ta:
    ta_sheet.cell(row=m["row"], column=2).value = f"Mock Property {m['row']}"
    ta_sheet.cell(row=m["row"], column=start_col).value = m["out"]
    ta_sheet.cell(row=m["row"], column=start_col+1).value = m["tgt"]
    ta_sheet.cell(row=m["row"], column=start_col+2).value = m["w1"]
    ta_sheet.cell(row=m["row"], column=start_col+3).value = m["w2"]
    ta_sheet.cell(row=m["row"], column=start_col+4).value = m["w3"]
    ta_sheet.cell(row=m["row"], column=start_col+5).value = m["w4"]
    ta_sheet.cell(row=m["row"], column=start_col+6).value = m["exp"]

# 2. MGMT FEE - Create new sheet
if "Mgmt Fee" in wb.sheetnames:
    del wb["Mgmt Fee"]
mgmt_sheet = wb.create_sheet("Mgmt Fee")
mgmt_headers = ["Property", "0-30", "30-60", "60-90", "90-120", "Target", "Received", "Expected"]
mgmt_sheet.append(mgmt_headers)
mgmt_sheet.append(["Hosa Road", 50000, 20000, 10000, 0, 100000, 20000, 150000])
mgmt_sheet.append(["Infantry", 0, 0, 50000, 100000, 150000, 0, 50000])
mgmt_sheet.append(["Koramangala", 10000, 0, 0, 15000, 30000, 50000, 0])
mgmt_sheet.append(["Whitefield", 0, 250000, 0, 0, 250000, 10000, 0])

# 3. PROFIT INCENTIVE - Append columns
pi_sheet = wb["Profit Incentive"]
pi_start_col = 25
pi_headers = ["KPI7", "0-30", "30-60", "60-90", "90-120", "Target", "Received", "Expected"]
for i, h in enumerate(pi_headers):
    pi_sheet.cell(row=1, column=pi_start_col + i).value = h

pi_mock = [
    {"row": 2, "030": 0, "3060": 50000, "6090": 0, "90120": 0, "tgt": 50000, "rcv": 10000, "exp": 0},
    {"row": 3, "030": 100000, "3060": 0, "6090": 150000, "90120": 200000, "tgt": 450000, "rcv": 0, "exp": 200000},
    {"row": 4, "030": 0, "3060": 0, "6090": 0, "90120": 500000, "tgt": 500000, "rcv": 50000, "exp": 0},
]

for m in pi_mock:
    pi_sheet.cell(row=m["row"], column=pi_start_col+1).value = m["030"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+2).value = m["3060"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+3).value = m["6090"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+4).value = m["90120"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+5).value = m["tgt"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+6).value = m["rcv"]
    pi_sheet.cell(row=m["row"], column=pi_start_col+7).value = m["exp"]

wb.save(EXCEL_PATH)
print("Data successfully injected.")
